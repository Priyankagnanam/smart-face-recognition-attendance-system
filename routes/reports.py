import logging
import csv
import io
from datetime import datetime, date, timedelta
from flask import Blueprint, render_template, request, jsonify, send_file, Response
from flask_login import login_required
from models.student import Student
from models.attendance import Attendance
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)
reports_bp = Blueprint('reports', __name__, url_prefix='/reports')


@reports_bp.route('/')
@login_required
def index():
    return render_template('reports.html')


@reports_bp.route('/api/data')
@login_required
def api_data():
    report_type = request.args.get('type', 'daily')
    date_str = request.args.get('date', '')
    department = request.args.get('department', '')

    today = date.today()

    if report_type == 'daily':
        query_date = today
        if date_str:
            try:
                query_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        records = get_attendance_for_date(query_date, department)
        return jsonify({
            'type': 'daily',
            'date': query_date.isoformat(),
            'records': records,
            'total': len(records),
            'present': sum(1 for r in records if r['status'] == 'Present'),
        })

    elif report_type == 'weekly':
        end_date = today
        if date_str:
            try:
                end_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        start_date = end_date - timedelta(days=7)
        records = get_attendance_for_range(start_date, end_date, department)
        return jsonify({
            'type': 'weekly',
            'from': start_date.isoformat(),
            'to': end_date.isoformat(),
            'records': records,
            'total': len(records),
        })

    elif report_type == 'monthly':
        year = today.year
        month = today.month
        if date_str:
            try:
                d = datetime.strptime(date_str, '%Y-%m').date()
                year, month = d.year, d.month
            except ValueError:
                try:
                    d = datetime.strptime(date_str, '%Y-%m-%d').date()
                    year, month = d.year, d.month
                except ValueError:
                    pass
        records = get_attendance_for_month(year, month, department)
        return jsonify({
            'type': 'monthly',
            'month': f'{year}-{month:02d}',
            'records': records,
            'total': len(records),
        })

    elif report_type == 'department':
        records = get_department_report(department)
        return jsonify({
            'type': 'department',
            'records': records,
            'total': len(records),
        })

    return jsonify({'records': [], 'total': 0})


def get_attendance_for_date(query_date, department=''):
    query = Attendance.query.filter(Attendance.attendance_date == query_date)
    if department:
        query = query.join(Student).filter(Student.department == department)
    records = query.order_by(Attendance.check_in_time.desc()).all()
    return format_records(records)


def get_attendance_for_range(start_date, end_date, department=''):
    query = Attendance.query.filter(
        Attendance.attendance_date >= start_date,
        Attendance.attendance_date <= end_date,
    )
    if department:
        query = query.join(Student).filter(Student.department == department)
    records = query.order_by(Attendance.attendance_date.desc(), Attendance.check_in_time.desc()).all()
    return format_records(records)


def get_attendance_for_month(year, month, department=''):
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    return get_attendance_for_range(start_date, end_date, department)


def get_department_report(department=''):
    if department:
        students = Student.query.filter_by(department=department).all()
    else:
        students = Student.query.all()

    report = []
    for student in students:
        total = Attendance.query.filter_by(student_id=student.student_id).count()
        report.append({
            'student_id': student.student_id,
            'name': student.name,
            'department': student.department,
            'year': student.year,
            'section': student.section,
            'total_attendance': total,
        })
    return report


def format_records(records):
    result = []
    for rec in records:
        student = Student.query.get(rec.student_id)
        result.append({
            'id': rec.id,
            'student_id': rec.student_id,
            'name': student.name if student else 'Unknown',
            'department': student.department if student else '',
            'year': student.year if student else '',
            'section': student.section if student else '',
            'date': rec.attendance_date.isoformat(),
            'time': rec.check_in_time.strftime('%H:%M:%S'),
            'status': rec.status,
            'confidence': round(rec.confidence_score * 100, 1) if rec.confidence_score else 0,
        })
    return result


def export_columns(records, report_type):
    """Return (headers, rows) for CSV/Excel export, normalizing the department
    report (which has total_attendance instead of per-day fields)."""
    if report_type == 'department':
        headers = ['Student ID', 'Name', 'Department', 'Year', 'Section', 'Total Attendance']
        rows = [[r['student_id'], r['name'], r['department'], r['year'], r['section'],
                 r['total_attendance']] for r in records]
        return headers, rows

    headers = ['Student ID', 'Name', 'Department', 'Year', 'Section', 'Date', 'Time', 'Status', 'Confidence (%)']
    rows = [[r['student_id'], r['name'], r['department'], r['year'], r['section'],
             r['date'], r['time'], r['status'], r['confidence']] for r in records]
    return headers, rows


@reports_bp.route('/api/export/csv')
@login_required
def export_csv():
    report_type = request.args.get('type', 'daily')
    date_str = request.args.get('date', '')
    department = request.args.get('department', '')
    today = date.today()

    records = []
    filename = f'report_{report_type}_{today.isoformat()}.csv'

    if report_type == 'daily':
        query_date = today
        if date_str:
            try:
                query_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        records = get_attendance_for_date(query_date, department)
    elif report_type == 'weekly':
        end_date = today
        if date_str:
            try:
                end_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        records = get_attendance_for_range(end_date - timedelta(days=6), end_date, department)
    elif report_type == 'monthly':
        year, month = today.year, today.month
        if date_str:
            try:
                d = datetime.strptime(date_str, '%Y-%m').date()
                year, month = d.year, d.month
            except ValueError:
                pass
        records = get_attendance_for_month(year, month, department)
    elif report_type == 'department':
        records = get_department_report(department)

    output = io.StringIO()
    writer = csv.writer(output)
    headers, rows = export_columns(records, report_type)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@reports_bp.route('/api/export/excel')
@login_required
def export_excel():
    report_type = request.args.get('type', 'daily')
    date_str = request.args.get('date', '')
    department = request.args.get('department', '')

    today = date.today()
    records = []
    filename = f'report_{report_type}_{today.isoformat()}.xlsx'

    if report_type == 'daily':
        query_date = today
        if date_str:
            try:
                query_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        records = get_attendance_for_date(query_date, department)
    elif report_type == 'weekly':
        end_date = today
        if date_str:
            try:
                end_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        records = get_attendance_for_range(end_date - timedelta(days=6), end_date, department)
    elif report_type == 'monthly':
        year, month = today.year, today.month
        if date_str:
            try:
                d = datetime.strptime(date_str, '%Y-%m').date()
                year, month = d.year, d.month
            except ValueError:
                pass
        records = get_attendance_for_month(year, month, department)
    elif report_type == 'department':
        records = get_department_report(department)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    headers, rows = export_columns(records, report_type)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
